# _*_ coding: utf_8  _*_
 
import openpyxl
import os
import re
import logging
import pprint

classReg = re.compile(r'\d{7}') # classReg = re.compile(r'\d{7}[zZ]?')  # classReg = re.compile('\d*7')

logging.disable(logging.CRITICAL)
# logging.basicConfig( filename='loglearn.txt',level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )                   
logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.critical('--------Start of program---------')

studentNo = input('please input the students No: ')
if( studentNo == '' ):
    studentNo = '201510040101'
    print(' '*2+studentNo)    
logging.info('studentNo is:%s', studentNo)

className = studentNo[2:6]+studentNo[7:10]
logging.info('className is:%s',' '*2+className)

# The student's marks need to write in the cell of the excel
# List is reasonable at here.
studentMarks = [
    ['课程',],
    ['学号', ],
    ['姓名', ],
    ['课堂平时成绩',],
    ['课堂期末成绩',],
    ['课堂总成绩',],
    ['实践成绩',],
    ['实验成绩',],
    ['总成绩',],
    ]
# Read marks
for twoDigit in range(33):
    studentNum = str(int(studentNo) + twoDigit)
    for fileNames in os.listdir('d:\\_PythonWorks\\execlOperate\\2016201701cj'):
        logging.debug(fileNames)
        classSearch = classReg.search(fileNames) # a = re.findall(classReg,fileNames)
        if classSearch:
            logging.debug(classSearch.group())
            logging.debug(className)
            if classSearch.group() == className:
                logging.debug(fileNames )
           
                wb = openpyxl.load_workbook( fileNames ,'r' )
                sheet = wb.get_sheet_by_name('page 1')

                for row in range(1,130):
                    logging.info('row is:%d',row)
                    logging.info( 'error test' + str( sheet['B'+str(row)].value ) )
                    if( sheet['B'+str(row)].value == int(studentNum) ):               #  学号在B列
                        logging.debug( ' '*2+str(row)+'  '+sheet['D'+str(row)].value )       #  姓名在D列
                        logging.debug(' 课堂平时成绩：'+ str( sheet['J'+str(row)].value ) )  #  课堂平时成绩在J列
                        logging.debug(' 课堂期末成绩：'+ str( sheet['M'+str(row)].value ) )  #  课堂期末成绩在M列
                        logging.debug(' 课堂总成绩：'+ str( sheet['O'+str(row)].value ) )    #  课堂总成绩在O列
                        logging.debug(' 实践成绩：'+ str( sheet['Q'+str(row)].value ) )      #  实践成绩在Q列
                        logging.debug(' 实验成绩：'+ str( sheet['R'+str(row)].value ) )      #  实验成绩在R列
                        logging.debug(' 总成绩：'+ str( sheet['S'+str(row)].value ) )        #  总成绩在S列
                        studentMarks[0].append(fileNames)
                        studentMarks[1].append(studentNum)
                        studentMarks[2].append(sheet['D'+str(row)].value)
                        studentMarks[3].append(sheet['J'+str(row)].value)
                        studentMarks[4].append(sheet['M'+str(row)].value)
                        studentMarks[5].append(sheet['O'+str(row)].value)
                        studentMarks[6].append(sheet['Q'+str(row)].value)
                        studentMarks[7].append(sheet['R'+str(row)].value)
                        studentMarks[8].append(sheet['S'+str(row)].value)                  
                        break
                    logging.error( 'error test')

# Write marks    
wb = openpyxl.Workbook()
wb.create_sheet(index=0,title=className)
sheet = wb.get_sheet_by_name(className)
col = 2       
for val in studentMarks:
    logging.info(val)
    for n in range(len(studentMarks[0])):         
        logging.info(val[n])
        sheet.cell(row = 2 + n,column = col).value = val[n]
    col +=1
newExcelName = className + '.xlsx'
wb.save(newExcelName)

print('Done!')

"""
201510040101

"""
